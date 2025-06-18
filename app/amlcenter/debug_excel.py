#!/usr/bin/env python3
import sys
import json
from openpyxl import load_workbook
import warnings

# Suppress warnings
warnings.filterwarnings("ignore", category=UserWarning)

def debug_excel_data(file_path):
    """Debug the Excel file to see exactly what data is being extracted"""
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        print(f"Available sheets: {workbook.sheetnames}")
        
        # Focus on BAA sheet (which represents BANK AL AMAL)
        sheet_name = "BAA"
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet {sheet_name} not found. Available sheets: {workbook.sheetnames}")
            return
            
        sheet = workbook[sheet_name]
        print(f"\n=== DEBUGGING SHEET: {sheet_name} ===")
        
        # Check the risk assessment area (rows 8-27)
        print("\nRisk assessment area (rows 8-27):")
        print("Row | Column A (Category)        | Column B (Factor)          | Column C (Empty)           | Column D (Profile)         | Column E (Rating)")
        print("----|---------------------------|----------------------------|----------------------------|----------------------------|------------------")
        
        for row in range(8, 28):
            try:
                cell_a = sheet.cell(row=row, column=1).value  # Column A
                cell_b = sheet.cell(row=row, column=2).value  # Column B  
                cell_c = sheet.cell(row=row, column=3).value  # Column C
                cell_d = sheet.cell(row=row, column=4).value  # Column D
                cell_e = sheet.cell(row=row, column=5).value  # Column E
                
                # Format for display
                a_str = str(cell_a)[:25] if cell_a else ""
                b_str = str(cell_b)[:25] if cell_b else ""
                c_str = str(cell_c)[:25] if cell_c else ""
                d_str = str(cell_d)[:25] if cell_d else ""
                e_str = str(cell_e)[:15] if cell_e else ""
                
                print(f"{row:2d}  | {a_str:<25} | {b_str:<25} | {c_str:<25} | {d_str:<25} | {e_str:<15}")
                
            except Exception as e:
                print(f"Error reading row {row}: {e}")
        
        # Check merged cells
        print(f"\nMerged ranges in {sheet_name}:")
        for merged_range in sheet.merged_cells.ranges:
            top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            print(f"  {merged_range}: '{top_left_cell.value}'")
        
        # Focus on specific rows that should have profile data
        print(f"\nDetailed analysis of key rows:")
        key_rows = [9, 10, 11, 12, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25]
        
        for row in key_rows:
            try:
                cell_a = sheet.cell(row=row, column=1).value
                cell_b = sheet.cell(row=row, column=2).value
                cell_c = sheet.cell(row=row, column=3).value
                cell_d = sheet.cell(row=row, column=4).value
                cell_e = sheet.cell(row=row, column=5).value
                
                print(f"\nRow {row}:")
                print(f"  A{row}: '{cell_a}'")
                print(f"  B{row}: '{cell_b}'")
                print(f"  C{row}: '{cell_c}'")
                print(f"  D{row}: '{cell_d}'")
                print(f"  E{row}: '{cell_e}'")
                
            except Exception as e:
                print(f"Error reading detailed row {row}: {e}")
                
    except Exception as e:
        print(f"Error loading workbook: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) > 1:
        debug_excel_data(sys.argv[1])
    else:
        debug_excel_data("uploads/excelFile-1749855242288.xlsx")
