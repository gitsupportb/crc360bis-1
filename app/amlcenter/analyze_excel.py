#!/usr/bin/env python3
"""
Analyze the Excel file structure to understand the exact format
"""

import openpyxl

def analyze_excel():
    wb = openpyxl.load_workbook('uploads/excelFile-1749850385370.xlsx')
    
    print("Available sheets:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        if 'AMAL' in sheet_name or 'POPULAIRE' in sheet_name or 'MAROC' in sheet_name:
            print(f"\n=== ANALYZING SHEET: {sheet_name} ===")
            sheet = wb[sheet_name]
            
            # Show the risk assessment area (rows 8-27) with detailed column analysis
            print("\nRisk assessment area (rows 8-27):")
            print("Row | Column A (Category)        | Column B (Factor)          | Column C (Empty)           | Column D (Profile)         | Column E (Rating)")
            print("----|---------------------------|----------------------------|----------------------------|----------------------------|------------------")
            for row in range(8, 28):
                row_data = []
                for col in range(1, 6):
                    cell = sheet.cell(row=row, column=col)
                    value = str(cell.value) if cell.value is not None else ""
                    row_data.append(f"{value[:26]:<26}")
                print(f"{row:2d}  | {' | '.join(row_data)}")

            # Also show which cells have actual data vs merged cells
            print("\nDetailed cell analysis:")
            for row in range(8, 28):
                for col in range(1, 6):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        col_letter = chr(64 + col)  # A, B, C, D, E
                        print(f"  {col_letter}{row}: '{cell.value}'")
                print()
            
            # Show merged ranges
            print(f"\nMerged ranges in {sheet_name}:")
            for merged_range in sheet.merged_cells.ranges:
                if merged_range.min_row >= 8 and merged_range.min_row <= 27:
                    top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    print(f"  {merged_range}: '{top_left_cell.value}'")
            
            break  # Analyze just the first relevant sheet for now

if __name__ == "__main__":
    analyze_excel()
