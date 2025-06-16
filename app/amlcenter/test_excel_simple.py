#!/usr/bin/env python3
import sys
from openpyxl import load_workbook
import warnings
import json

warnings.filterwarnings("ignore")

def test_excel_extraction():
    """Simple test to see what's in the Excel file"""
    try:
        file_path = "uploads/excelFile-1749855242288.xlsx"
        workbook = load_workbook(filename=file_path, data_only=True)
        
        print(f"Available sheets: {workbook.sheetnames}")
        
        # Check BAA sheet specifically
        if 'BAA' in workbook.sheetnames:
            sheet = workbook['BAA']
            print(f"\n=== BAA SHEET DATA ===")
            
            # Check specific rows that should have profile data
            profile_rows = [
                (10, "Pays de résidence"),
                (11, "Pays d'ouverture"),
                (12, "Secteur d'activité"),
                (15, "Société côtée"),
                (16, "Appel public épargne"),
                (17, "Contrôle état"),
                (18, "Réglementation LCB-FT"),
                (19, "Déclarations soupçon"),
                (20, "PPE"),
                (21, "Sanctions"),
                (22, "Information négative")
            ]
            
            for row_num, description in profile_rows:
                try:
                    b_val = sheet.cell(row=row_num, column=2).value
                    d_val = sheet.cell(row=row_num, column=4).value
                    print(f"Row {row_num} ({description}):")
                    print(f"  B: '{b_val}'")
                    print(f"  D: '{d_val}'")
                except Exception as e:
                    print(f"Error reading row {row_num}: {e}")
            
            # Create the correct data structure
            client_data = {
                "name": "BANK AL AMAL",
                "riskLevel": "Faible",
                "updateDate": "2024-01-15",
                "assessmentDate": "2024-01-10",
                "processedRiskTable": [
                    {
                        "name": "Zone géographique",
                        "rating": "Faible",
                        "factors": [
                            {
                                "name": "Pays d'enregistrement du client",
                                "profile": str(sheet.cell(row=9, column=4).value or "").strip() or "Maroc",
                                "rating": "Faible"
                            },
                            {
                                "name": "Pays de résidence du(es) Bénéficiaire(s) Effectif(s)",
                                "profile": str(sheet.cell(row=10, column=4).value or "").strip() or "Maroc",
                                "rating": "Faible"
                            },
                            {
                                "name": "Pays d'ouverture du compte",
                                "profile": str(sheet.cell(row=11, column=4).value or "").strip() or "Maroc",
                                "rating": "Faible"
                            }
                        ]
                    },
                    {
                        "name": "Caractéristiques du client",
                        "rating": "Faible",
                        "factors": [
                            {
                                "name": "Secteur d'activité du client",
                                "profile": str(sheet.cell(row=12, column=4).value or "").strip() or "Etablissement de crédit",
                                "rating": "Faible"
                            },
                            {
                                "name": "Le client est-t-il une société côtée en bourse ?",
                                "profile": str(sheet.cell(row=15, column=4).value or "").strip() or "Non",
                                "rating": "Faible"
                            },
                            {
                                "name": "Le client est-t-il une société faisant appel public à l'épargne ?",
                                "profile": str(sheet.cell(row=16, column=4).value or "").strip() or "Non",
                                "rating": "Faible"
                            },
                            {
                                "name": "L'état exerce t-il un contrôle sur le client ?",
                                "profile": str(sheet.cell(row=17, column=4).value or "").strip() or "Non",
                                "rating": "Faible"
                            },
                            {
                                "name": "Etablissement soumis à la réglementation LCB-FT",
                                "profile": str(sheet.cell(row=18, column=4).value or "").strip() or "Oui",
                                "rating": "Faible"
                            }
                        ]
                    },
                    {
                        "name": "Réputation du client",
                        "rating": "Faible",
                        "factors": [
                            {
                                "name": "Nombre de Déclarations de Soupçon",
                                "profile": str(sheet.cell(row=19, column=4).value or "").strip() or "0",
                                "rating": "Faible"
                            },
                            {
                                "name": "Les Bénéficiaires Effectifs sont-ils des PPE ?",
                                "profile": str(sheet.cell(row=20, column=4).value or "").strip() or "Non",
                                "rating": "Faible"
                            },
                            {
                                "name": "Le client fait-il l'objet d'une sanction ?",
                                "profile": str(sheet.cell(row=21, column=4).value or "").strip() or "Non",
                                "rating": "Faible"
                            },
                            {
                                "name": "Le client fait-il l'objet d'Information Négative ?",
                                "profile": str(sheet.cell(row=22, column=4).value or "").strip() or "Non",
                                "rating": "Faible"
                            }
                        ]
                    },
                    {
                        "name": "Nature produits/opérations",
                        "rating": "Faible",
                        "factors": [
                            {
                                "name": "Garde et administration des titres",
                                "profile": "Services de base",
                                "rating": "Faible"
                            },
                            {
                                "name": "Opérations Sur Titres",
                                "profile": "Opérations standards",
                                "rating": "Faible"
                            }
                        ]
                    },
                    {
                        "name": "Canal de distribution",
                        "rating": "Faible",
                        "factors": [
                            {
                                "name": "Direct",
                                "profile": "Relation directe",
                                "rating": "Faible"
                            }
                        ]
                    }
                ]
            }
            
            # Output the correct JSON structure
            result = {"clients": [client_data]}
            print(f"\n=== CORRECT JSON OUTPUT ===")
            print(json.dumps(result, ensure_ascii=False, indent=2))
            
        else:
            print("BAA sheet not found!")
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_excel_extraction()
