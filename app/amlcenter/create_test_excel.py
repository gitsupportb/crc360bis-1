#!/usr/bin/env python3
"""
Create a test Excel file with the exact structure expected by the system
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import datetime

def create_test_excel():
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RED MED ASSET MANAGEMENT"
    
    # Set up the header information
    ws['A1'] = 'BANK AL AMAL'
    ws['H1'] = datetime.date(2023, 10, 14)  # Date de MAJ
    ws['H3'] = datetime.date(2023, 2, 10)   # Date d'EER
    
    # Set up the table headers
    ws['A6'] = 'Évaluation des risques BC/FT'
    ws['A7'] = 'Identification des risques BC/FT'
    
    # Column headers
    ws['A8'] = 'Facteurs de risques'
    ws['B8'] = 'Profil de risques'
    ws['E8'] = 'Notation de risque'
    
    # Zone géographique category (rows 9-11)
    ws['A9'] = 'Zone géographique'
    ws['E9'] = 'Faible'
    
    ws['A10'] = "Pays d'enregistrement du client"
    ws['B10'] = 'Maroc'
    ws['E10'] = 'Faible'
    
    ws['A11'] = "Pays de résidence du(es) Bénéficiaire(s) Effectif(s)"
    ws['B11'] = 'Maroc'
    ws['E11'] = 'Faible'
    
    # Caractéristiques du client category (rows 12-18)
    ws['A12'] = 'Caractéristiques du client'
    ws['E12'] = 'Faible'
    
    ws['A13'] = "Secteur d'activité du client"
    ws['B13'] = 'Société de gestion des OPCVM'
    ws['E13'] = 'Faible'
    
    ws['A14'] = "Chiffre d'Affaires du client"
    ws['B14'] = ''
    ws['E14'] = 'Faible'
    
    ws['A15'] = "Date de création de la personne morale"
    ws['B15'] = 'Établissement de crédit'
    ws['E15'] = 'Faible'
    
    ws['A16'] = "Le client est-il une société cotée en bourse ?"
    ws['B16'] = 'Non'
    ws['E16'] = 'Faible'
    
    ws['A17'] = "Le client est-il une société faisant appel public à l'épargne ?"
    ws['B17'] = 'Non'
    ws['E17'] = 'Faible'
    
    ws['A18'] = "L'état exerce t-il un contrôle sur le client ?"
    ws['B18'] = 'Non'
    ws['E18'] = 'Faible'
    
    # Réputation du client category (rows 19-22)
    ws['A19'] = 'Réputation du client'
    ws['E19'] = 'Faible'
    
    ws['A20'] = "Nombre de Déclarations de Soupçon"
    ws['B20'] = '0'
    ws['E20'] = 'Faible'
    
    ws['A21'] = "Les Bénéficiaires Effectifs sont-ils des PPE ?"
    ws['B21'] = 'Non'
    ws['E21'] = 'Faible'
    
    ws['A22'] = "Le client fait-il l'objet d'une sanction, ou s'il est activé dans un pays sous embargo ?"
    ws['B22'] = 'Non'
    ws['E22'] = 'Faible'
    
    # Nature produits/opérations category (rows 23-24)
    ws['A23'] = 'Nature produits/opérations'
    ws['E23'] = 'Faible'
    
    ws['A24'] = "Garde et administration des titres"
    ws['B24'] = ''
    ws['E24'] = 'Faible'
    
    # Canal de distribution category (row 25)
    ws['A25'] = 'Canal de distribution'
    ws['E25'] = 'Faible'
    
    ws['A26'] = "Direct"
    ws['B26'] = ''
    ws['E26'] = 'Faible'
    
    # Overall risk level (row 27)
    ws['A27'] = 'Niveau risque'
    ws['B27'] = 'Faible'
    
    # Save the file
    filename = 'test_risk_assessment.xlsx'
    wb.save(filename)
    print(f"Created test Excel file: {filename}")
    return filename

if __name__ == "__main__":
    create_test_excel()
