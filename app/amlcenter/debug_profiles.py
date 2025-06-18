#!/usr/bin/env python3
"""
Debug profile extraction
"""

import openpyxl

def debug_profiles():
    wb = openpyxl.load_workbook('uploads/excelFile-1749849010039.xlsx')
    sheet = wb['RED MED ASSET MANAGEMENT']

    # Get merged cell ranges
    merged_ranges = {}
    for merged_range in sheet.merged_cells.ranges:
        top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_ranges[(row, col)] = top_left_cell.value

    def get_cell_value(row, col):
        if (row, col) in merged_ranges:
            return merged_ranges[(row, col)]
        else:
            cell = sheet.cell(row=row, column=col)
            return cell.value

    # Check problematic rows
    for row in [13, 14, 23, 24, 25]:
        print(f'Row {row}:')
        for col in range(1, 6):
            value = get_cell_value(row, col)
            print(f'  Column {col}: "{value}"')
        
        # Check what would be extracted as profile
        profile = ''
        profile_value = get_cell_value(row, 4)  # Column D
        if profile_value is not None and str(profile_value).strip() and str(profile_value).strip() != 'None':
            cell_value = str(profile_value).strip()
            if cell_value not in ['Faible', 'Moyen', 'Élevé', 'Elevé']:
                profile = cell_value
        
        if not profile:
            profile_value = get_cell_value(row, 3)  # Column C
            if profile_value is not None and str(profile_value).strip() and str(profile_value).strip() != 'None':
                cell_value = str(profile_value).strip()
                if cell_value not in ['Faible', 'Moyen', 'Élevé', 'Elevé']:
                    profile = cell_value

        if not profile:
            profile = 'Non spécifié'
            
        print(f'  Extracted profile: "{profile}"')
        print()

if __name__ == "__main__":
    debug_profiles()
