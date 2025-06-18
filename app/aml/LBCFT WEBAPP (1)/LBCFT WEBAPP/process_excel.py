import pandas as pd
import json
import sys
import os
import datetime
from openpyxl import load_workbook

# Define known categories for better detection
# Export as a module variable so it can be imported by server.js
KNOWN_CATEGORIES = [
    'Zone géographique',
    'Caractéristiques du client',
    'Réputation du client',
    'Nature produits/opérations',
    'Canal de distribution'
]

# Make categories available for import
if __name__ != "__main__":
    # When imported as a module
    __all__ = ['KNOWN_CATEGORIES', 'process_excel_file', 'extract_specific_range', 'process_risk_table']

def extract_specific_range(workbook, sheet_name, range_str):
    """Extract data from a specific range in an Excel sheet"""
    # Parse the range string (e.g., 'A8:E25')
    try:
        # Get the sheet object
        sheet = workbook[sheet_name]
        
        # Parse the range string
        start, end = range_str.split(':')
        
        # Convert Excel column letters to indices (0-based)
        start_col = ord(start[0]) - ord('A')
        start_row = int(start[1:]) - 1  # Convert to 0-based index
        
        end_col = ord(end[0]) - ord('A')
        end_row = int(end[1:]) - 1  # Convert to 0-based index
        
        # Initialize the result array
        range_data = []
        
        # First pass: identify all categories and their positions
        category_positions = []
        for row in range(start_row, end_row + 1):
            # Check if this row contains a category name (in column A)
            cell_value = sheet.cell(row=row+1, column=1).value
            if cell_value:
                # More strict category detection
                is_category = False
                cell_str = str(cell_value).strip()
                for cat in KNOWN_CATEGORIES:
                    if (cell_str == cat or 
                        cell_str.startswith(cat) or 
                        cat.lower() in cell_str.lower() or
                        cell_str.lower() in cat.lower()):
                        is_category = True
                        break
                
                if is_category:
                    category_positions.append({
                        'row': row,
                        'name': cell_value
                    })
        
        # If no categories found, try a more lenient approach with partial matching
        if not category_positions:
            print(f"No exact category matches found, trying partial matching for sheet {sheet_name}", file=sys.stderr)
            for row in range(start_row, end_row + 1):
                cell_value = sheet.cell(row=row+1, column=1).value
                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    # Check for partial matches with category keywords
                    if any(keyword in cell_str for keyword in ['zone', 'géo', 'client', 'caractéristique', 'réputation', 'produit', 'opération', 'canal', 'distribution']):
                        category_positions.append({
                            'row': row,
                            'name': cell_value
                        })
                        print(f"Found partial category match: '{cell_value}' in sheet {sheet_name}", file=sys.stderr)
        
        # If still no categories found, check for any bold text or formatting that might indicate a category
        if not category_positions:
            print(f"No partial category matches found, checking for formatting indicators in sheet {sheet_name}", file=sys.stderr)
            for row in range(start_row, end_row + 1):
                cell_value = sheet.cell(row=row+1, column=1).value
                if cell_value and not sheet.cell(row=row+1, column=2).value:
                    # If column A has a value but column B is empty, it might be a category header
                    category_positions.append({
                        'row': row,
                        'name': cell_value
                    })
                    print(f"Possible category header found by format: '{cell_value}' in sheet {sheet_name}", file=sys.stderr)
        
        # If still no categories found, create a default category
        if not category_positions:
            print(f"Warning: No categories found in range {range_str} for sheet {sheet_name}", file=sys.stderr)
            # Create a default category at the start of the range
            category_positions.append({
                'row': start_row,
                'name': 'Données non catégorisées'
            })
        
        # Second pass: extract all data with proper category assignment
        for i, category_pos in enumerate(category_positions):
            next_category_pos = category_positions[i + 1] if i + 1 < len(category_positions) else None
            category_end_row = next_category_pos['row'] - 1 if next_category_pos else end_row
            
            # Get category rating from column E
            rating_cell_value = sheet.cell(row=category_pos['row']+1, column=5).value
            category_rating = 'Faible'
            if rating_cell_value:
                rating_str = str(rating_cell_value).strip()
                if rating_str in ['Faible', 'Moyen', 'Élevé', 'Elevé']:
                    category_rating = rating_str
                    # Normalize 'Elevé' to 'Élevé'
                    if category_rating == 'Elevé':
                        category_rating = 'Élevé'
            
            # Add the category row
            category_row_data = {'A': category_pos['name'], 'isCategory': True, 'rating': category_rating}
            range_data.append(category_row_data)
            
            # Add all risk factors for this category
            for row in range(category_pos['row'] + 1, category_end_row + 1):
                row_data = {'category': category_pos['name']}
                has_data = False
                
                # Extract all columns for this row
                for col in range(start_col, end_col + 1):
                    col_letter = chr(65 + col)
                    cell_value = sheet.cell(row=row+1, column=col+1).value
                    
                    if cell_value is not None:
                        # Convert to string and strip whitespace
                        if isinstance(cell_value, (str, int, float)):
                            row_data[col_letter] = str(cell_value).strip()
                        else:
                            row_data[col_letter] = cell_value
                        has_data = True
                    else:
                        row_data[col_letter] = None
                
                # Determine rating for this factor
                if row_data.get('E'):
                    rating_str = str(row_data['E']).strip()
                    if rating_str in ['Faible', 'Moyen', 'Élevé', 'Elevé']:
                        row_data['rating'] = rating_str
                        # Normalize 'Elevé' to 'Élevé'
                        if row_data['rating'] == 'Elevé':
                            row_data['rating'] = 'Élevé'
                    else:
                        row_data['rating'] = category_rating
                else:
                    row_data['rating'] = category_rating
                
                # Only add rows that have data in column A (factor name)
                if has_data and row_data.get('A'):
                    range_data.append(row_data)
        
        # Debug output
        print(f"Extracted {len(range_data)} rows from range {range_str} in sheet {sheet_name}", file=sys.stderr)
        
        return range_data
        
    except Exception as e:
        print(f"Error extracting range {range_str} from sheet {sheet_name}: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return []

def process_risk_table(range_data):
    """Process the extracted data into a structured format for display"""
    processed_data = []
    current_category = None
    
    # Handle empty or None range_data
    if not range_data:
        print("Warning: No range data provided to process_risk_table", file=sys.stderr)
        return [{
            'name': 'Données non disponibles',
            'rating': 'Faible',
            'factors': [{
                'name': 'Information manquante',
                'profile': 'Aucune donnée trouvée dans la plage spécifiée',
                'rating': 'Faible'
            }]
        }]
    
    try:
        # First pass: identify all categories and create their objects
        for row in range_data:
            if not row:
                continue
                
            # Check if this is a category header
            if row.get('isCategory'):
                # Normalize rating value
                rating = row.get('rating', 'Faible')
                if rating == 'Elevé':
                    rating = 'Élevé'
                    
                # Create a new category object
                current_category = {
                    'name': row.get('A', 'Catégorie non spécifiée'),
                    'rating': rating,
                    'factors': []
                }
                
                # Add to processed data
                processed_data.append(current_category)
                print(f"Added category: {current_category['name']} with rating: {rating}", file=sys.stderr)
        
        # If no categories were found, create a default one
        if not processed_data:
            processed_data.append({
                'name': 'Données non catégorisées',
                'rating': 'Faible',
                'factors': []
            })
            print("No categories found, created default category", file=sys.stderr)
        
        # Second pass: assign risk factors to their categories
        for row in range_data:
            if not row or row.get('isCategory'):
                continue
                
            # Find the category this factor belongs to
            category_name = row.get('category')
            category = None
            
            # Find the matching category
            for cat in processed_data:
                if cat['name'] == category_name:
                    category = cat
                    break
            
            # If category not found, use the first one as fallback
            if not category and processed_data:
                category = processed_data[0]
                print(f"Warning: Category '{category_name}' not found for factor '{row.get('A')}', using '{category['name']}' instead", file=sys.stderr)
            
            if category and row.get('A'):
                # Combine columns B, C, and D for the profile
                profile_parts = []
                for col in ['B', 'C', 'D']:
                    if row.get(col) is not None:
                        try:
                            # Convert to string and strip whitespace
                            value = str(row.get(col)).strip()
                            # Only exclude values that are risk ratings or empty
                            if value and value not in ['Faible', 'Moyen', 'Élevé', 'Elevé', 'None', 'nan']:
                                # Check if it's a numeric value that might be a date
                                try:
                                    float_val = float(value)
                                    # Skip if it looks like a numeric code or date
                                    if float_val > 10000:
                                        continue
                                except ValueError:
                                    # Not a number, include it
                                    pass
                                profile_parts.append(value)
                        except Exception as e:
                            # Log the error but continue
                            print(f"Error processing profile part: {str(e)}", file=sys.stderr)
                            pass
                
                profile = ' '.join(profile_parts)
                # If profile is empty, provide a default
                if not profile.strip():
                    profile = 'Non spécifié'
                
                # Normalize rating value
                rating = row.get('rating', category['rating'])
                if rating == 'Elevé':
                    rating = 'Élevé'
                
                # Add this factor to the category
                category['factors'].append({
                    'name': row.get('A', 'Facteur non spécifié'),
                    'profile': profile or 'Non spécifié',
                    'rating': rating
                })
                print(f"Added factor: {row.get('A')} to category: {category['name']}", file=sys.stderr)
        
        # Ensure each category has at least one factor
        for category in processed_data:
            if not category['factors']:
                category['factors'].append({
                    'name': 'Information non disponible',
                    'profile': 'Aucune donnée trouvée pour cette catégorie',
                    'rating': category['rating']
                })
                print(f"Added default factor to empty category: {category['name']}", file=sys.stderr)
        
        # Ensure all categories have proper ratings
        for category in processed_data:
            # If any factor has a higher risk rating than the category, update the category rating
            highest_rating = category['rating']
            for factor in category['factors']:
                if factor['rating'] == 'Élevé' and highest_rating != 'Élevé':
                    highest_rating = 'Élevé'
                elif factor['rating'] == 'Moyen' and highest_rating == 'Faible':
                    highest_rating = 'Moyen'
            
            # Update category rating if needed
            if highest_rating != category['rating']:
                print(f"Updating category {category['name']} rating from {category['rating']} to {highest_rating} based on factor ratings", file=sys.stderr)
                category['rating'] = highest_rating
        
        # Debug output
        print(f"Processed {len(processed_data)} categories with a total of {sum(len(cat.get('factors', [])) for cat in processed_data)} factors", file=sys.stderr)
        
    except Exception as e:
        # Log the error but return what we have so far
        import traceback
        print(f"Error processing risk table: {str(e)}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        
        # If we have no processed data yet, add a default category
        if not processed_data:
            processed_data.append({
                'name': 'Erreur de traitement',
                'rating': 'Faible',
                'factors': [{
                    'name': 'Erreur',
                    'profile': f'Une erreur est survenue lors du traitement: {str(e)}',
                    'rating': 'Faible'
                }]
            })
    
    return processed_data

def extract_client_info(workbook, sheet_name):
    """Extract client information from the Excel sheet"""
    sheet = workbook[sheet_name]
    
    # Check if sheet is empty or has no data
    # Ensure max_row and max_column are at least 1 to avoid index errors
    if not hasattr(sheet, 'max_row') or not hasattr(sheet, 'max_column') or sheet.max_row < 1 or sheet.max_column < 1:
        print(f"Warning: Sheet {sheet_name} appears to be empty or invalid", file=sys.stderr)
        return {
            'name': sheet_name,
            'riskLevel': 'Faible',
            'updateDate': '',
            'assessmentDate': ''
        }
    
    # Extract client name (usually in cell C1 or similar)
    client_name = sheet_name
    # First try to find client name in specific cells that typically contain it
    for row in range(1, min(10, sheet.max_row + 1)):  # Check first few rows with safety check
        try:
            # Check cells in columns 1-5 for client name
            for col in range(1, 6):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and any(name in cell_value for name in ['RED MED', 'BANK AL AMAL', 'SECURITIES', 'CLIENT', 'CUSTOMER']):
                    client_name = cell_value
                    break
        except ValueError as e:
            print(f"Warning: Error accessing cell at row {row}, column 3: {str(e)}", file=sys.stderr)
            continue
    
    # Extract risk level (usually at the bottom of the sheet)
    risk_level = 'Faible'  # Default value
    if sheet.max_row > 0:
        # Ensure we don't go below row 1
        start_row = max(1, sheet.max_row - 14)
        for row in range(sheet.max_row, start_row, -1):  # Check last few rows with safety check
            try:
                if sheet.cell(row=row, column=2).value == 'Niveau risque' and sheet.cell(row=row, column=6).value:
                    risk_level = sheet.cell(row=row, column=6).value
                    break
            except ValueError as e:
                print(f"Warning: Error accessing cell for risk level at row {row}: {str(e)}", file=sys.stderr)
                continue
    
    # Extract dates
    update_date = ''
    assessment_date = ''
    for row in range(1, min(sheet.max_row + 1, 50)):  # Limit to first 50 rows for efficiency
        try:
            if sheet.cell(row=row, column=7).value == 'Date de MAJ' and sheet.cell(row=row, column=8).value:
                date_value = sheet.cell(row=row, column=8).value
                # Format date if it's a datetime object
                if isinstance(date_value, (datetime.datetime, datetime.date)):
                    update_date = date_value.strftime('%Y-%m-%d')
                # Handle Excel serial date numbers
                elif isinstance(date_value, (int, float)):
                    try:
                        # Convert Excel serial date to datetime
                        # Excel dates are number of days since 1900-01-01, with a leap year bug
                        date_obj = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=date_value)
                        update_date = date_obj.strftime('%Y-%m-%d')
                    except Exception:
                        update_date = str(date_value)
                else:
                    update_date = str(date_value)
            if sheet.cell(row=row, column=7).value == "Date d'EER" and sheet.cell(row=row, column=8).value:
                date_value = sheet.cell(row=row, column=8).value
                # Format date if it's a datetime object
                if isinstance(date_value, (datetime.datetime, datetime.date)):
                    assessment_date = date_value.strftime('%Y-%m-%d')
                # Handle Excel serial date numbers
                elif isinstance(date_value, (int, float)):
                    try:
                        # Convert Excel serial date to datetime
                        date_obj = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=date_value)
                        assessment_date = date_obj.strftime('%Y-%m-%d')
                    except Exception:
                        assessment_date = str(date_value)
                else:
                    assessment_date = str(date_value)
        except ValueError as e:
            print(f"Warning: Error accessing cell for dates at row {row}: {str(e)}", file=sys.stderr)
            continue
    
    return {
        'name': client_name,
        'riskLevel': risk_level,
        'updateDate': update_date,
        'assessmentDate': assessment_date
    }

def process_excel_file(file_path):
    """Process the Excel file and return structured data"""
    # Load the workbook with warnings suppressed
    import warnings
    import traceback
    
    # Suppress the specific Data Validation warning
    warnings.filterwarnings("ignore", category=UserWarning, 
                          message="Data Validation extension is not supported and will be removed")
    
    # Load the workbook
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {str(e)}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        return {'clients': []}
    
    clients = []
    
    # List of sheets to skip (typically non-client sheets or special sheets)
    skip_sheets = ['Instructions', 'Guide', 'Template', 'Index', 'Profil de risque']
    
    # Process each sheet in the workbook (each sheet represents a client)
    for sheet_name in workbook.sheetnames:
        try:
            # Skip known non-client sheets
            if sheet_name in skip_sheets:
                print(f"Skipping non-client sheet: {sheet_name}", file=sys.stderr)
                continue
                
            # Skip sheets with insufficient data
            if workbook[sheet_name].max_row < 10:
                print(f"Skipping sheet with insufficient data: {sheet_name}", file=sys.stderr)
                continue
            
            print(f"Processing sheet: {sheet_name}", file=sys.stderr)
            
            # Extract client information with error handling
            try:
                client_info = extract_client_info(workbook, sheet_name)
            except Exception as e:
                print(f"Error extracting client info from sheet {sheet_name}: {str(e)}", file=sys.stderr)
                # Use default values if extraction fails
                client_info = {
                    'name': sheet_name,
                    'riskLevel': 'Faible',
                    'updateDate': '',
                    'assessmentDate': ''
                }
            
            # Extract data from the specific range A8:E25 with improved error handling
            try:
                # Get the sheet object and verify it has enough rows and columns before attempting to extract the range
                sheet = workbook[sheet_name]
                if sheet.max_row >= 8 and sheet.max_column >= 5:
                    range_data = extract_specific_range(workbook, sheet_name, 'A8:E25')
                    print(f"Successfully extracted {len(range_data)} rows from range A8:E25 in sheet {sheet_name}", file=sys.stderr)
                else:
                    print(f"Sheet {sheet_name} does not have sufficient rows/columns for range A8:E25, using empty data", file=sys.stderr)
                    range_data = []
            except Exception as e:
                print(f"Error extracting range data from sheet {sheet_name}: {str(e)}", file=sys.stderr)
                traceback.print_exc(file=sys.stderr)
                range_data = []
            
            # Process the risk table with error handling
            try:
                processed_risk_table = process_risk_table(range_data)
                print(f"Successfully processed risk table with {len(processed_risk_table)} categories for sheet {sheet_name}", file=sys.stderr)
            except Exception as e:
                print(f"Error processing risk table for sheet {sheet_name}: {str(e)}", file=sys.stderr)
                traceback.print_exc(file=sys.stderr)
                processed_risk_table = [{
                    'name': 'Erreur de traitement',
                    'rating': 'Faible',
                    'factors': [{
                        'name': 'Erreur',
                        'profile': f'Une erreur est survenue lors du traitement: {str(e)}',
                        'rating': 'Faible'
                    }]
                }]
            
            # Create client object
            client = {
                'name': client_info['name'],
                'riskLevel': client_info['riskLevel'],
                'updateDate': client_info['updateDate'],
                'assessmentDate': client_info['assessmentDate'],
                'processedRiskTable': processed_risk_table,
                'extractedRiskData': range_data,  # Include raw extracted data for fallback
                'knownCategories': KNOWN_CATEGORIES  # Add known categories for reference in template
            }
            
            clients.append(client)
            
            print(f"Processed client: {client['name']} with {len(processed_risk_table)} risk categories", file=sys.stderr)
            
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {str(e)}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
    
    return {'clients': clients}

def main():
    """Main function to process Excel file"""
    if len(sys.argv) < 2:
        print("Error: Please provide the path to the Excel file")
        sys.exit(1)
    
    file_path = sys.argv[1]
    # Ensure the file path is properly formatted and exists
    if not os.path.isabs(file_path):
        # Convert to absolute path if it's not already
        file_path = os.path.abspath(file_path)
    
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} does not exist")
        sys.exit(1)
    
    # Process the Excel file
    try:
        result = process_excel_file(file_path)
        # Output the result as JSON
        print(json.dumps(result))
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()